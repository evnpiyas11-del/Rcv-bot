import logging
import os
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
import openpyxl

# --- কনফিগারেশন ---
TOKEN = '8796638911:AAHQODPuXyvJHnXN28WQjJDJKI3qwsTicFs'
ADMIN_IDS = [7229266894, 8060113003] # আপনারা দুইজন

# আপনার দেওয়া পাবলিক গ্রুপের ইউজারনেম
RECEIVED_GROUP_USERNAME = '@rcvid11' 
REJECTED_GROUP_USERNAME = '@reject1234'

# লগিং সেটআপ
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# মেইন মেনু কিবোর্ড (টাইপিং বক্সের নিচে থাকবে)
def get_main_reply_keyboard():
    keyboard = [
        [KeyboardButton("📁 File Submit")],
        [KeyboardButton("❓ কিভাবে সাবমিট করবো")],
        [KeyboardButton("📞 Support")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        'কি আবস্থা মামা! নিচের বাটন থেকে একটি অপশন সিলেক্ট করো:', 
        reply_markup=get_main_reply_keyboard()
    )

async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "📁 File Submit":
        keyboard = [[InlineKeyboardButton("Facebook", callback_data='fb')],
                    [InlineKeyboardButton("Instagram", callback_data='ig')]]
        await update.message.reply_text("🔖 মামা কোন ধরনের ফাইল সাবমিট করবা?", reply_markup=InlineKeyboardMarkup(keyboard))
    elif text == "📞 Support":
        await update.message.reply_text("📞 সাপোর্টের জন্য যোগাযোগ করুন: https://t.me/iklash11")
    elif text == "❓ কিভাবে সাবমিট করবো":
        await update.message.reply_text("📂 ফাইল সাবমিট করার ভিডিও লিংক ➡️ [আপনার ভিডিওর লিংক এখানে]")

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    # অ্যাডমিন অ্যাকশন (রিসিভ/রিজেক্ট)
    if data.startswith('admin_'):
        parts = data.split('_')
        action = parts[1]
        user_id = int(parts[2])
        file_msg_id = int(parts[3])
        admin_who_clicked = query.from_user.username if query.from_user.username else query.from_user.first_name

        original_caption = query.message.caption

        if action == 'receive':
            # ১. ইউজারকে সেই ফাইলের ওপর রিপ্লাই দিয়ে জানানো
            try:
                await context.bot.send_message(chat_id=user_id, text="✅ আপনার এই ফাইলটি রিসিপ করা হয়েছে।", reply_to_message_id=file_msg_id)
            except: pass
            
            # ২. অ্যাডমিন মেসেজের স্ট্যাটাস আপডেট
            await query.edit_message_caption(caption=original_caption + f"\n\nStatus: RECEIVED ✅\nBy Admin: @{admin_who_clicked}")
            
            # ৩. রিসিভ গ্রুপে ফাইল পাঠানো
            try:
                await context.bot.send_document(
                    chat_id=RECEIVED_GROUP_USERNAME,
                    document=query.message.document.file_id,
                    caption=f"✅ RECEIVED RECORD\n\n{original_caption}\n\nProcessed By: @{admin_who_clicked}"
                )
            except Exception as e:
                print(f"Receive group error: {e}")

        elif action == 'reject':
            # ১. ইউজারকে সেই ফাইলের ওপর রিপ্লাই দিয়ে জানানো
            try:
                await context.bot.send_message(chat_id=user_id, text="❌ আপনার এই ফাইলটি রিজেক্ট করা হয়েছে। 📞 বিস্তারিত জানতে @iklash11", reply_to_message_id=file_msg_id)
            except: pass
            
            # ২. অ্যাডমিন মেসেজের স্ট্যাটাস আপডেট
            await query.edit_message_caption(caption=original_caption + f"\n\nStatus: REJECTED ❌\nBy Admin: @{admin_who_clicked}")
            
            # ৩. রিজেক্ট গ্রুপে ফাইল পাঠানো
            try:
                await context.bot.send_document(
                    chat_id=REJECTED_GROUP_USERNAME,
                    document=query.message.document.file_id,
                    caption=f"❌ REJECTED RECORD\n\n{original_caption}\n\nProcessed By: @{admin_who_clicked}"
                )
            except Exception as e:
                print(f"Reject group error: {e}")

    # ক্যাটাগরি সিলেকশন লজিক
    elif data == 'fb':
        keyboard = [[InlineKeyboardButton("Facebook Cookies", callback_data='cat_FB-Cookies')],
                    [InlineKeyboardButton("Facebook 2FA", callback_data='cat_FB-2FA')],
                    [InlineKeyboardButton("Facebook Hotmail", callback_data='cat_FB-Hotmail')]]
        await query.edit_message_text("Facebook ক্যাটাগরি সিলেক্ট করুন:", reply_markup=InlineKeyboardMarkup(keyboard))
    elif data == 'ig':
        keyboard = [[InlineKeyboardButton("Instagram 2FA", callback_data='cat_IG-2FA')],
                    [InlineKeyboardButton("Instagram Cookies", callback_data='cat_IG-Cookies')]]
        await query.edit_message_text(" 🔍 Instagram ক্যাটাগরি সিলেক্ট করুন:", reply_markup=InlineKeyboardMarkup(keyboard))
    elif data.startswith('cat_'):
        category = data.split('_')[1]
        context.user_data['selected_category'] = category
        await query.edit_message_text(f"📄 আপনার {category} Excel ফাইলটি পাঠান।")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if 'selected_category' not in context.user_data:
        await update.message.reply_text("আগে ক্যাটাগরি সিলেক্ট করুন!")
        return

    user = update.message.from_user
    category = context.user_data['selected_category']
    file_id = update.message.document.file_id
    file_name = update.message.document.file_name
    file_msg_id = update.message.message_id

    # এক্সেল ফাইল থেকে আইডি কাউন্ট
    try:
        new_file = await update.message.document.get_file()
        await new_file.download_to_drive(file_name)
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        id_count = sheet.max_row if sheet.max_row > 0 else 0
        wb.close()
        os.remove(file_name)
    except:
        id_count = "N/A"

    # ইউজারকে কনফার্মেশন ও কিবোর্ড ফিরিয়ে দেওয়া
    await update.message.reply_text(
        f"✅ আপনার {id_count}টি আইডি সাবমিট হয়েছে ✔️। এডমিন চেক করে রিপ্লে দিবে...",
        reply_markup=get_main_reply_keyboard()
    )

    # অ্যাডমিনদের কাছে ফাইল পাঠানো
    admin_caption = (
        f"📩 নতুন ফাইল এসেছে!\n\n"
        f"📜 ক্যাটাগরি: {category}\n"
        f"🔢 আইডি সংখ্যা: {id_count} টি\n"
        f"🆔 ইউজার : @{user.username if user.username else 'ইউজারনেম নেই'}"
    )
    
    admin_keyboard = [[InlineKeyboardButton("✅ Receive", callback_data=f"admin_receive_{user.id}_{file_msg_id}"),
                       InlineKeyboardButton("❌ Reject", callback_data=f"admin_reject_{user.id}_{file_msg_id}")]]

    for admin in ADMIN_IDS:
        try:
            await context.bot.send_document(chat_id=admin, document=file_id, caption=admin_caption, reply_markup=InlineKeyboardMarkup(admin_keyboard))
        except: continue

if __name__ == '__main__':
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(filters.TEXT & (~filters.COMMAND), message_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    print("বট চলছে মামা... গ্রুপ ফরওয়ার্ডিং ফুল সেটআপ করা হয়েছে।")
    app.run_polling()
